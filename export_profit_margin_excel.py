#export_profit_margin_excel.py

import os
import sys
import pandas as pd
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

def export_profit_pool_to_excel(
    profit_pool,
    end_date,
    template_path="export_cdp_poc_stock_60D_sort_temple.xlsm",
):
    """将多周期利润率池数据（profit_margin_pool）无缝写入Excel模板。
    
    股票代码和股票名称输出时自动添加双引号。
    数据统一从第二列（B列）开始对齐写入，从第2行开始（行2输出日期，行3输出表头，行4及以后输出数据）。
    """
    if not profit_pool:
        print("提示：profit_pool 数据为空，未执行写入。")
        return

    # 1. 定义统一的表格表头（严格对应您的利润率池字典键名）
    headers = [
        "股票代码",
        "股票名称",
        "5天短线利润",
        "60日线(季度)利润",
        "120日线(半年)利润"
    ]

    # 2. 将利润池数据转换为二维列表
    table_rows = []
    for item in profit_pool:
        # 💡 数据装载时，使用 f-string 强制为代码和名称包裹双引号
        stock_code_with_quotes = f'"{item.get("股票代码", "")}"'
        stock_name_with_quotes = f'"{item.get("股票名称", "")}"'
        
        table_rows.append([
            stock_code_with_quotes,
            stock_name_with_quotes,
            item.get("5天短线利润", "0.0%"),
            item.get("60日线(季度)利润", "0.0%"),
            item.get("120日线(半年)利润", "0.0%")
        ])

    # 转为 DataFrame
    excel_data = pd.DataFrame(table_rows, columns=headers)

    # ======================
    # 写入Excel模板
    # ======================

    # 加载模板，保留 VBA 宏
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["Sheet1"]

    # 💡 核心控制：从B列开始
    START_COL = 2     # B列
    START_ROW = 2     # 从第2行开始写入

    # 【第一行】：输出结束日期 end_date 
    ws.cell(row=START_ROW, column=START_COL, value=end_date)
    START_ROW += 1

    # 【第二行】：输出字段表头
    for col_idx, header in enumerate(excel_data.columns, start=START_COL):
        ws.cell(row=START_ROW, column=col_idx, value=header)
    START_ROW += 1

    # 【第三行及以后】：依次无缝输出每只股票的具体多周期利润率数据
    for row_offset, row in enumerate(excel_data.itertuples(index=False), start=START_ROW):
        for col_idx, value in enumerate(row, start=START_COL):
            ws.cell(row=row_offset, column=col_idx, value=value)

    # ======================
    # 动态路径生成与保存
    # ======================

    # 1. 定义目标文件夹路径
    output_dir = Path(end_date)

    # 2. 自动创建文件夹
    output_dir.mkdir(parents=True, exist_ok=True)

    # 3. 拼接完整的文件路径
    # 生成特定文件名的宏模板输出文件
    excel_path = output_dir / f"选股池_多周期利润率_{end_date}.xlsm"

    # 保存文件
    wb.save(excel_path)
    print(f"{excel_path} 利润率池数据（已加双引号）保存完成")


# ==========================================
# 🚀 自动化测试入口
# ==========================================
if __name__ == "__main__":
    print("=== 开始执行利润率 Excel 导出功能本地测试 ===")

    # 1. 模拟排序后的真实数据输入（注意：传入前可用上一步的 .sort() 排好序）
    mock_profit_pool = [
        {
            "股票代码": "688635.SS",
            "股票名称": "长进电子",
            "5天短线利润": "12.34%",
            "60日线(季度)利润": "45.67%",
            "120日线(半年)利润": "89.01%"
        },
        {
            "股票代码": "603228.SH",
            "股票名称": "昼马科技",
            "5天短线利润": "-1.23%",
            "60日线(季度)利润": "15.45%",
            "120日线(半年)利润": "-3.12%"
        }
    ]

    # 2. 排序测试（按您上一轮要求的60日线降序逻辑）
    mock_profit_pool.sort(
        key=lambda x: float(x["60日线(季度)利润"].replace("%", "")), 
        reverse=True
    )

    # 3. 指定参数
    test_date = "2026-07-28"
    test_template = "export_Profit_temple.xlsm"  # 💡 请确保本地有这个模板，或者改成您实际的模板名

    # 4. 执行导出
    export_profit_pool_to_excel(
        profit_pool=mock_profit_pool,
        end_date=test_date,
        template_path=test_template
    )
    
    print("=== 测试流运行完毕 ===")
