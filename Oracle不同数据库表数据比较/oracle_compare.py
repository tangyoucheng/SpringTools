import os
import sys
import pandas as pd
import oracledb
import math
from sqlalchemy import create_engine, text
from openpyxl.styles import PatternFill

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from common_oracle_config import URL_OLD, URL_NEW
from common_oracle_meta import get_oracle_meta

# SQLAlchemy の Engine オブジェクトを作成（これで UserWarning が完全に消えます）
engine_old = create_engine(URL_OLD)
engine_new = create_engine(URL_NEW)

# 対象のテーブル名（Oracleはデフォルト大文字のため、必ず大文字で指定）
#TABLE_NAME = "TCMMSGINFO" 
TABLE_NAME = "TCMMTHOJO" 
# 【重要】業務主キーを指定（単一、または複合主キーのリスト）
# 例1（単一カラム）: BUSINESS_PKS = ["ORDER_NO"]
# 例2（複合カラム）: BUSINESS_PKS = ["USER_ID", "PRODUCT_ID"]
#BUSINESS_PKS = ["SQSEQ"] 
BUSINESS_PKS = ["CDCMP", "KBMST", "CDCODE"]
# 範囲検索（BETWEEN）に使用する基準カラム（通常は上記の主キーのいずれか1つを指定）
#RANGE_COL = "SQSEQ"
RANGE_COL = "CDCMP"
# ==================== ★ 比較対象の複数テーブル設定 ★ ====================
# ここに対象のテーブル、業務主キーのリストをすべて定義します。
# 何表あっても、ここに追記するだけで自動で順番に処理されます。
#
# 【重要】業務主キーを指定（単一、または複合主キーのリスト）
# 例1（単一カラム）: BUSINESS_PKS = ["ORDER_NO"]
# 例2（複合カラム）: BUSINESS_PKS = ["USER_ID", "PRODUCT_ID"]
TABLES_CONFIG = [
    {
        "TABLE_NAME": "TCMMSGINFO",
        "BUSINESS_PKS": ["SQSEQ"]
    },
    {
        "TABLE_NAME": "TCMMTHOJO",
        "BUSINESS_PKS": ["CDCMP", "KBMST", "CDCODE"]
    }
]

# ==================== ★ Excel出力の配置設定 ★ ====================
# Excelの出力開始位置を指定します（0から始まるインデックス）
# 例: START_ROW = 2  -> 3行目から開始（上側に2行空ける）
# 例: START_COL = 1  -> B列から開始（左側に1列空ける）
START_ROW = 2  
START_COL = 1  

# ==================== 3. ページングを使用した分割比較 ====================
def execute_compare_for_table(target_config):
    table_name = target_config["TABLE_NAME"]
    business_pks = target_config["BUSINESS_PKS"]
    
    print(f"\n==================================================")
    print(f" テーブル 【{table_name}】 の比較処理を開始します...")
    print(f"==================================================")
    
    print("Oracleのテーブル定義と論理名（コメント）を取得中...")
    type_dict, comment_dict = get_oracle_meta(engine_new, table_name)
    
    # 複合キーと単一キーの双方で安全に動作するハッシュ関数用の文字列表現を作成
    pk_concat_str = " || ',' || ".join(business_pks)
    
    # ORA_HASHを用いてデータを100個のバケットに均等分割（千万級データ対応）
    total_chunks = 100  
    all_diff_rows = []
    diff_cells_coord = []
    current_row_idx = 0  
    is_first_diff = True
    
    with engine_old.connect() as conn_old, engine_new.connect() as conn_new:
        for bucket_id in range(total_chunks):
            
            # 新環境からバケットに該当するデータを取得
            sql_new = f"""
                SELECT * FROM {table_name}
                WHERE ORA_HASH({pk_concat_str}, {total_chunks - 1}) = :bucket_id
                ORDER BY {', '.join(business_pks)}
            """
            df_new = pd.read_sql_query(text(sql_new), conn_new, params={"bucket_id": bucket_id})
            df_new.columns = df_new.columns.str.upper()
            
            if df_new.empty:
                continue
                
            # 旧环境也同步读取相同哈希桶内的数据，确保数据范围完全对称一致
            sql_old = f"""
                SELECT * FROM {table_name}
                WHERE ORA_HASH({pk_concat_str}, {total_chunks - 1}) = :bucket_id
                ORDER BY {', '.join(business_pks)}
            """
            df_old = pd.read_sql_query(text(sql_old), conn_old, params={"bucket_id": bucket_id})
            df_old.columns = df_old.columns.str.upper()

            # 業務主キーをインデックスに設定
            df_new.set_index(business_pks, inplace=True)
            df_old.set_index(business_pks, inplace=True)
            
            common_ids = df_old.index.intersection(df_new.index)
            if common_ids.empty:
                continue
                
            sub_old = df_old.loc[common_ids]
            sub_new = df_new.loc[common_ids]
            
            for idx in common_ids:
                row_old = sub_old.loc[idx]
                row_new = sub_new.loc[idx]
                
                if not row_old.fillna('').equals(row_new.fillna('')):
                    
                    # 最初の1回目のみ、最上部にフィールド英名とデータ型を出力
                    if is_first_diff:
                        # 1行目：フィールド物理名
                        row_eng = {col: col for col in sub_old.columns}
                        for k in business_pks: row_eng[k] = k
                        row_eng['比較タイプ'] = 'フィールド物理名(英名)'
                        all_diff_rows.append(row_eng)
                        current_row_idx += 1
                        
                        # 2行目：データ型と長さ
                        row_type = {col: type_dict.get(col, '') for col in sub_old.columns}
                        for k in business_pks: row_type[k] = type_dict.get(k, '')
                        row_type['比較タイプ'] = 'データ型(長さ)'
                        all_diff_rows.append(row_type)
                        current_row_idx += 1
                        
                        is_first_diff = False  
                    
                    row_new_data = row_new.to_dict()
                    row_old_data = row_old.to_dict()
                    row_new_data['比較タイプ'] = '新データ (NEW)'
                    row_old_data['比較タイプ'] = '旧データ (OLD)'
                    
                    # 【型エラー・複数キー完全対応版】
                    # 単一キー（文字列）と複合キー（タプル）を判定して安全にマッピング
                    if len(business_pks) == 1:
                        pk_key = business_pks[0]
                        row_new_data[pk_key] = idx
                        row_old_data[pk_key] = idx
                    else:
                        for k, v in zip(business_pks, idx):
                            row_new_data[k] = v
                            row_old_data[k] = v
                    
                    # 差分セルの座標記録
                    new_row_pos = current_row_idx  
                    for col in sub_old.columns:
                        val_old = str(row_old[col]).strip() if pd.notnull(row_old[col]) else ''
                        val_new = str(row_new[col]).strip() if pd.notnull(row_new[col]) else ''
                        if val_old != val_new:
                            diff_cells_coord.append((new_row_pos, col))
                    
                    all_diff_rows.append(row_new_data)
                    all_diff_rows.append(row_old_data)
                    all_diff_rows.append({'比較タイプ': '-----------------------------------'})
                    
                    current_row_idx += 3  
                    
            if (bucket_id + 1) % 10 == 0:
                print(f"進捗: 進捗ブロック {bucket_id + 1} / {total_chunks} の比較が完了...")
                
    # ==================== 4. Excelファイルへの出力と色付け調整 ====================
    if all_diff_rows:
        result_df = pd.DataFrame(all_diff_rows)
        cols = ['比較タイプ'] + business_pks + [c for c in result_df.columns if c not in ['比較タイプ'] + business_pks]
        result_df = result_df[cols]
        
        display_comment_dict = comment_dict.copy()
        display_comment_dict['比較タイプ'] = '論理名'  
        for pk in business_pks:
            display_comment_dict[pk] = comment_dict.get(pk, pk)
            
        original_cols = list(result_df.columns)
        result_df.rename(columns=display_comment_dict, inplace=True)
        
        output_file = f"{table_name}_diff_result.xlsx"
        
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            result_df.to_excel(writer, index=False, sheet_name='Data_Diff', startrow=START_ROW, startcol=START_COL)
            
            workbook = writer.book
            worksheet = writer.sheets['Data_Diff']
            pink_fill = PatternFill(start_color='FFD2D2', end_color='FFD2D2', fill_type='solid')
            
            for row_pos, col_name in diff_cells_coord:
                col_idx = original_cols.index(col_name) + START_COL + 1 
                excel_row = row_pos + START_ROW + 2  
                
                cell = worksheet.cell(row=excel_row, column=col_idx)
                cell.fill = pink_fill
                
        print(f"🎉 【{table_name}】 比較完了！保存先: {output_file}")
    else:
        print(f" 比較完了：【{table_name}】 に差分データはありませんでした。")

# ==================== 4. メイン実行部 ====================
if __name__ == "__main__":
    # TABLES_CONFIG に定義されたテーブルをループで1表ずつ自動実行
    for config in TABLES_CONFIG:
        execute_compare_for_table(config)
    print("\n🚀 すべてのテーブルの比較処理が正常に終了しました！")
