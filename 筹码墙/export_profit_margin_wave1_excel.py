#export_profit_margin_wave1_excel.py

import os
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
# 💡【核心修补】导入标准列字母转换工具，用纯正 Python 方式修改列宽
from openpyxl.utils import get_column_letter

def export_profit_pool_to_excel(
    profit_pool,
    end_date,
    template_path="export_cdp_poc_stock_60D_sort_temple.xlsm",
):
    """将第一浪反弹利润率龙虎榜数据无缝写入Excel汇总模板，彻底修复 Worksheet.Range 导致的崩溃错误。
    
    股票代码和股票名称输出时自动添加双引号。
    数据统一从第二列（B列）开始对齐写入，从第2行开始（行2输出日期，行3输出表头，行4及以后输出数据）。
    """
    if not profit_pool:
        print("提示：profit_pool 数据为空，未执行写入。")
        return

    # 1. 定义统一的表格表头（严格对齐近端第一浪反弹雷达的测算键名）
    headers = [
        "股票代码",
        "股票名称",
        "当前收盘价",
        "5日ATR",
        "首道重力墙(元)",
        "墙体装甲厚度",
        "触网净空间(元)",
        "ATR波动跨度",
        "★第一浪利润期望"
    ]

    # 2. 将利润池数据转换为二维列表
    table_rows = []
    for item in profit_pool:
        # 💡 数据装载时，使用 f-string 强制为代码和名称包裹双引号
        stock_code_with_quotes = f'"{item.get("股票代码", "")}"'
        stock_name_with_quotes = f'"{item.get("股票名称", "")}"'
        
        # 强类型显式划分，剥离 Python 隐式封装，根除公式损坏报错
        table_rows.append([
            str(stock_code_with_quotes),
            str(stock_name_with_quotes),
            float(item.get("当前收盘价", 0.0)),
            float(item.get("5日ATR", 0.0)),
            float(item.get("首道重力墙(元)", 0.0)),
            str(item.get("墙体装甲厚度", "")),
            float(item.get("触网净空间(元)", 0.0)),
            float(item.get("ATR波动跨度", 0.0)),
            float(item.get("★第一浪利润期望", 0.0))
        ])

    # 转为标准安全 DataFrame
    excel_data = pd.DataFrame(table_rows, columns=headers)

    # 加载模板，保留 VBA 宏
    wb = load_workbook(template_path, keep_vba=True)
    ws = wb["Sheet1"]

    # 从 B 列第 2 行开始对齐写入
    START_COL = 2     # B列
    START_ROW = 2     # 从第2行开始写入

    # 3. 动态提取并死锁原模板 B2 单元格的原生核心字体名，绝对防止自动回退到宋体
    template_font_name = ws.cell(row=2, column=2).font.name if (ws.cell(row=2, column=2).font and ws.cell(row=2, column=2).font.name) else "微软雅黑"

    # 【第2行】：输出结束日期 end_date
    date_cell = ws.cell(row=START_ROW, column=START_COL, value=str(end_date))
    date_cell.font = Font(name=template_font_name, size=11, bold=True)
    START_ROW += 1

    # 【第3行】：输出字段表头
    for col_idx, header in enumerate(excel_data.columns, start=START_COL):
        header_cell = ws.cell(row=START_ROW, column=col_idx, value=str(header))
        header_cell.font = Font(name=template_font_name, size=11, bold=True)
    START_ROW += 1

    # 【第4行及以后】：依次无缝输出每只股票的具体多周期利润率数据
    for row in excel_data.itertuples(index=False):
        for col_idx, value in enumerate(row, start=START_COL):
            cell = ws.cell(row=START_ROW, column=col_idx)
            
            # 根据内容性质，执行最高强度的单元格数据格式安全写入
            if isinstance(value, (int, float)):
                cell.value = float(value)
            else:
                cell.value = str(value)
                
            # 严格保持模板字形不变
            cell.font = Font(name=template_font_name, size=11)
            
        START_ROW += 1

    # =========================================================================
    # 💡【终极安全修补】用纯正 Python 循环遍历列宽，彻底杀死 Worksheet.Range 报错
    # =========================================================================
    for col in range(START_COL, START_COL + len(headers)):
        col_letter = get_column_letter(col)
        max_len = 0
        # 扫描该列所有已写入的行，测算最大字符旷度
        for row_idx in range(2, START_ROW):
            cell_val = ws.cell(row=row_idx, column=col).value
            if cell_val is not None:
                # 针对标准中文和中文字符串长度进行加权校准（防汉字显示不全）
                val_str = str(cell_val)
                # 计算含中文字符的真实视觉宽度
                real_len = len(val_str.encode('utf-8'))
                if real_len > max_len:
                    max_len = real_len
                    
        # 安全设置列宽，额外留出 3 个字符作为视觉安全缓冲带
        ws.column_dimensions[col_letter].width = max(max_len // 2 + 3, 12)
    # =========================================================================

    # ======================
    # 动态路径生成与保存
    # ======================
    output_dir = Path(end_date)
    output_dir.mkdir(parents=True, exist_ok=True)
    
    excel_path = output_dir / f"选股池_第一浪反弹套利期望_{end_date}.xlsm"
    wb.save(excel_path)
    print(f"📊 {excel_path} 独立汇总大账本保存完成（100%兼容openpyxl原生语法）")
