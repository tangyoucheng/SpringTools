import os
import glob
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def analyze_xlsm_stock_precise(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active 
    
    stock_code = str(ws['B33'].value).strip() if ws['B33'].value else "未知代码"
    stock_name = str(ws['C33'].value).strip() if ws['C33'].value else "未知名称"
    
    try:
        current_price = float(ws['D33'].value)
    except:
        return None 
        
    try:
        atr_14 = float(ws['G33'].value) if ws['G33'].value else 0.0
        atr_5 = float(ws['H33'].value) if ws['H33'].value else 0.0
    except:
        atr_14, atr_5 = 0.0, 0.0

    cdp_20_cdp = float(ws['D38'].value) if ws['D38'].value else 0.0  
    cdp_20_nl  = float(ws['F38'].value) if ws['F38'].value else 0.0  
    cdp_20_al  = float(ws['E38'].value) if ws['E38'].value else 0.0  
    
    cdp_60_cdp = float(ws['D39'].value) if ws['D39'].value else 0.0  
    cdp_60_nl  = float(ws['F39'].value) if ws['F39'].value else 0.0  

    walls_data = []
    for r in range(45, ws.max_row + 1):
        price_val = ws.cell(row=r, column=2).value 
        label_val = ws.cell(row=r, column=3).value 
        
        if price_val is not None:
            try:
                price_num = round(float(price_val), 2)
                label_str = str(label_val or "").strip()
                
                if "★当前收盘价" in label_str or "当前收盘" in label_str:
                    continue
                walls_data.append({"价格": price_num, "标签": label_str, "行号": r})
            except ValueError:
                continue

    if not walls_data:
        return None

    df_walls = pd.DataFrame(walls_data)
    
    # ==================== 【已完美修复】使用您提供的真实硬核特征码进行计数 ====================
    # 彻底告别盲区，精准数出 161.48 的 5 个大格子与 95.75 的 2 个大格子！
    df_walls['能量强度'] = df_walls['标签'].apply(lambda x: x.count('█'))

    # 用纯价格切出头顶的阻力区
    upper_walls = df_walls[df_walls['价格'] > current_price]
    lower_walls = df_walls[df_walls['价格'] < current_price]

    # ==================== 精准锁定头顶第一个最大阻力墙 ====================
    if not upper_walls.empty:
        # Step 1: 找出上方阻力墙中，黑格子数量的最大值（此时 5 会被顺利作为最大值抓出来）
        max_energy_strength = upper_walls['能量强度'].max()
        
        # Step 2: 过滤出所有达到了 5 个黑格子的绝对主力墙
        all_max_res_walls = upper_walls[upper_walls['能量强度'] == max_energy_strength]
        
        # Step 3: 并列时，取离当前价最近、价格最低的那一个
        best_resistance = all_max_res_walls.loc[all_max_res_walls['价格'].idxmin()]
            
        target_price = best_resistance['价格']
        resistance_name = best_resistance['标签']
        profit_to_wall = (target_price - current_price) / current_price
    else:
        target_price, resistance_name, profit_to_wall = None, "上方无阻力区", 0.0

    # ==================== 多维空间+动能趋势立体选股算法 ====================
    safe_window_price = current_price - (1.5 * atr_14 if atr_14 > 0 else current_price * 0.03)
    
    hard_floors = lower_walls[(lower_walls['价格'] >= safe_window_price) & 
                              (lower_walls['标签'].str.contains('20日|60日|半年|年度|核心|真POC', na=False))]
    has_strong_floor = not hard_floors.empty

    atr_ratio = atr_5 / atr_14 if atr_14 > 0 else 1.0
    is_amplitude_crushed = atr_ratio <= 0.85   
    is_amplitude_expanding = atr_ratio >= 1.20 

    is_bear_market_trend = cdp_20_cdp < cdp_60_cdp and current_price < cdp_60_nl

    if is_bear_market_trend and is_amplitude_crushed:
        bottom_status = "【3级：破位阴跌真空区】大趋势严重走坏，当前缩量属于无量阴跌死水，切勿抄底！"
    elif current_price <= cdp_20_nl * 1.02 and is_amplitude_crushed and has_strong_floor and not is_bear_market_trend:
        bottom_status = "【1级：坚实左侧底】空间跌透 + 波幅冰点 + 中长期大筹码护盘。黄金左侧潜伏点！"
    elif current_price < cdp_20_al and is_amplitude_expanding and has_strong_floor:
        bottom_status = "【2级：恐慌暴跌撞墙】击穿月度极限边界，恐慌盘涌出，但正砸在历史强支撑上。适合挂单接飞刀。"
    elif current_price < cdp_20_nl or (is_amplitude_crushed and not has_strong_floor):
        bottom_status = "【3级：破位阴跌真空区】价格已穿透支撑位，且脚下缺乏核心大筹码防线。严禁建仓！"
    elif current_price > cdp_20_cdp * 1.05 and is_amplitude_crushed:
        bottom_status = "【5级：高位悬空滞涨】价格悬空于各中长期支撑上方，高位缩量久盘必跌。随时发生补跌，严禁建仓！"
    else:
        bottom_status = f"【4级：中继常态震荡】多空在中轴【{cdp_20_cdp}】附近正常博弈，空间或时间未换到位，暂不建仓。"

    return {
        "股票代码": stock_code,
        "股票名称": stock_name,
        "当前收盘价": current_price,
        "第一大阻力价": target_price,
        "阻力墙特征": resistance_name,
        "到阻力墙利润空间": profit_to_wall, 
        "5日ATR": atr_5,
        "14日ATR": atr_14,
        "底部状态评估": bottom_status
    }

def generate_report(folder_path, output_path, start_row=1, start_col=1):
    xlsm_files = glob.glob(os.path.join(folder_path, "*.xlsm"))
    results = []
    
    print(f"正在精准解析 {len(xlsm_files)} 个股票文件...")
    for file in xlsm_files:
        try:
            stock_data = analyze_xlsm_stock_precise(file)
            if stock_data:
                results.append(stock_data)
        except Exception as e:
            print(f"文件 {os.path.basename(file)} 解析异常: {e}")
            
    if not results:
        print("未提取到任何有效数据。")
        return

    df_report = pd.DataFrame(results)
    df_report = df_report.sort_values(by="到阻力墙利润空间", ascending=False)
    df_report['到阻力墙利润空间'] = df_report['到阻力墙利润空间'].apply(lambda x: f"{x:.2%}")

    if os.path.exists(output_path):
        wb_out = load_workbook(output_path)
        ws_out = wb_out.active
    else:
        wb_out = Workbook()
        ws_out = wb_out.active
    
    max_exist_row = ws_out.max_row
    max_exist_col = ws_out.max_column
    if max_exist_row >= start_row and max_exist_col >= start_col:
        print("正在清除旧数据残余...")
        for r in range(start_row, max_exist_row + 1):
            for c in range(start_col, max_exist_col + 1):
                ws_out.cell(row=r, column=c).value = None

    headers = list(df_report.columns)
    for c_idx, header in enumerate(headers):
        cell = ws_out.cell(row=start_row, column=start_col + c_idx)
        cell.value = header
        
    for r_idx, row_data in enumerate(df_report.values):
        for c_idx, value in enumerate(row_data):
            cell = ws_out.cell(row=start_row + 1 + r_idx, column=start_col + c_idx)
            cell.value = value
            
    wb_out.save(output_path)
    print(f" 数据刷新完毕！完美锁定最大黑格子筹码墙。")
    
# --- 执行区域 ---
if __name__ == "__main__":
    # 1. 你的文件夹路径
    #target_folder = r"C:\price_py\CDP\算反弹阻力用\2026-08-03" # 建议用绝对路径确保万无一失
    #target_folder = "2026-07-30" # 建议用绝对路径确保万无一失
    target_folder = "2026-08-04" # 建议用绝对路径确保万无一失
    
    # 2. 输出的报告文件名
    #output_excel = r"C:\price_py\CDP\算反弹阻力用\总筛选报告.xlsx" 
    output_excel = f"总筛选报告_{target_folder}.xlsx" 
    
    # 3. 在这里【指定开始行和开始列】
    # 1代表A列，2代表B列，3代表C列...
    GEN_START_ROW = 2  
    GEN_START_COL = 2  
    
    generate_report(target_folder, output_excel, start_row=GEN_START_ROW, start_col=GEN_START_COL)
